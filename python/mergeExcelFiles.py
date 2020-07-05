#!/usr/bin/env python3
import os
import openpyxl

def processFileData(excelFile, mergedData):
  wb = openpyxl.load_workbook(filename=excelFile)
  dataSheetName = wb.sheetnames[0]
  ws = wb[dataSheetName]

  rowNum = 0
  for rowData in ws.values:
    if rowNum == 0 :
      rowNum = rowNum + 1
      continue            # 跳过第一行 - 非数据的表头
    else:
      rowNum = rowNum + 1

    province = rowData[0] #假设第一列存的是 省份
    if province in mergedData:
      mergedData[province].append(rowData)
    else:
      mergedData[province] = [rowData]


def processExcelFiles(path, mergedData):
  files = os.listdir(path)
  print("开始分析目录 {p} 下的excel文件".format(p = path))
  for f in files:
      filePath = os.path.join(path, f)
      if os.path.isfile(filePath) and filePath.endswith(".xlsx"):
        processFileData(filePath, mergedData)


def saveProcessedData(provincesData, saveDir):
  if not os.path.isdir(saveDir):  # 保存目录不存在则自动创建
    os.mkdir(saveDir)

  print("将保存的省份文件个数：{c}".format(c=len(provincesData)))
  for province, rowsData in provincesData.items():
    saveFile = os.path.join(saveDir, str(province) + ".xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rowsData:
      ws.append(row)
    wb.save(saveFile)
  print("{c} 个处理后的文件已经保存到 {p}：".format(c=len(provincesData), p=saveDir))


if __name__ == "__main__":
  mergedData = {}     #内存字典变量，用省份作字典的key，用于管理分析excel得到数据
  processExcelFiles("./excel/", mergedData)
  saveProcessedData(mergedData, "./out/")
  